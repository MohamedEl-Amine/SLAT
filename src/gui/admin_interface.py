"""
Admin interface for SLAT - Configuration and management.
"""

from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, 
                             QTableWidget, QTableWidgetItem, QTabWidget, QCheckBox, QMessageBox, 
                             QTimeEdit, QComboBox, QHeaderView, QFileDialog, QDialog, QFormLayout,
                             QGroupBox, QTextEdit, QScrollArea, QFrame, QListWidget, QRadioButton,
                             QButtonGroup, QDateEdit, QSpinBox, QSplitter)
from PyQt5.QtCore import Qt, QTime, QDate
from PyQt5.QtGui import QFont, QPixmap, QImage, QColor, QBrush
import csv
import sqlite3
from datetime import datetime, timedelta
from io import BytesIO
import base64
import calendar

class EmployeeProfileDialog(QDialog):
    def __init__(self, db, employee_id):
        super().__init__()
        self.db = db
        self.employee_id = employee_id
        self.employee = self.db.get_employee(employee_id)
        
        if not self.employee:
            QMessageBox.critical(self, "Erreur", "Employé introuvable.")
            self.reject()
            return
            
        self.setWindowTitle(f"Profil employé - {self.employee.name}")
        self.setMinimumSize(700, 550)
        self.resize(900, 650)  # Initial size, but resizable
        
        self.setup_ui()
        self.load_employee_data()
        
    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # Employee Details Section
        details_group = QGroupBox("Détails employé")
        details_layout = QFormLayout()
        
        self.name_edit = QLineEdit()
        self.name_edit.setText(self.employee.name)
        details_layout.addRow("Nom :", self.name_edit)
        
        self.id_label = QLabel(self.employee.employee_id)
        details_layout.addRow("ID employé :", self.id_label)
        
        self.status_label = QLabel("Activé" if self.employee.enabled else "Désactivé")
        self.status_label.setStyleSheet("color: green;" if self.employee.enabled else "color: red;")
        details_layout.addRow("Statut :", self.status_label)
        
        self.created_label = QLabel(self.employee.created_at.strftime("%Y-%m-%d %H:%M") if self.employee.created_at else "N/A")
        details_layout.addRow("Créé :", self.created_label)
        
        details_group.setLayout(details_layout)
        layout.addWidget(details_group)
        
        # Identification Methods Section
        methods_group = QGroupBox("Méthodes d'identification")
        methods_layout = QVBoxLayout()
        
        # QR Code Section
        qr_layout = QHBoxLayout()
        qr_layout.addWidget(QLabel("Code QR :"))
        
        self.qr_status = QLabel("Non généré" if not self.employee.qr_code else "Généré")
        self.qr_status.setStyleSheet("color: red;" if not self.employee.qr_code else "color: green;")
        qr_layout.addWidget(self.qr_status)
        
        self.view_qr_btn = QPushButton("Voir QR")
        self.view_qr_btn.clicked.connect(self.view_qr)
        qr_layout.addWidget(self.view_qr_btn)
        
        methods_layout.addLayout(qr_layout)
        
        # Face Recognition Section
        face_layout = QHBoxLayout()
        face_layout.addWidget(QLabel("Reconnaissance faciale :"))
        
        self.face_status = QLabel("Non défini" if not self.employee.face_embedding else "Défini")
        self.face_status.setStyleSheet("color: red;" if not self.employee.face_embedding else "color: green;")
        face_layout.addWidget(self.face_status)
        
        self.set_face_btn = QPushButton("Définir visage")
        self.set_face_btn.clicked.connect(self.set_face)
        face_layout.addWidget(self.set_face_btn)
        
        methods_layout.addLayout(face_layout)
        
        methods_group.setLayout(methods_layout)
        layout.addWidget(methods_group)
        
        # Recent Attendance Section
        attendance_group = QGroupBox("Présences récentes (10 derniers enregistrements)")
        attendance_layout = QVBoxLayout()
        
        self.attendance_table = QTableWidget()
        self.attendance_table.setColumnCount(4)
        self.attendance_table.setHorizontalHeaderLabels(["Action", "Heure", "Méthode", "Appareil"])
        self.attendance_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.attendance_table.setEditTriggers(QTableWidget.NoEditTriggers)  # Make table readonly
        attendance_layout.addWidget(self.attendance_table)
        
        attendance_group.setLayout(attendance_layout)
        layout.addWidget(attendance_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.save_btn = QPushButton("Sauvegarder modifications")
        self.save_btn.clicked.connect(self.save_changes)
        button_layout.addWidget(self.save_btn)
        
        self.toggle_status_btn = QPushButton("Désactiver" if self.employee.enabled else "Activer")
        self.toggle_status_btn.clicked.connect(self.toggle_status)
        button_layout.addWidget(self.toggle_status_btn)
        
        close_btn = QPushButton("Fermer")
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)
        
    def load_employee_data(self):
        # Load recent attendance logs
        logs = self.db.get_employee_logs(self.employee_id, limit=10)
        self.attendance_table.setRowCount(len(logs))
        
        for row, log in enumerate(logs):
            self.attendance_table.setItem(row, 0, QTableWidgetItem(log.action))
            self.attendance_table.setItem(row, 1, QTableWidgetItem(log.timestamp.strftime("%Y-%m-%d %H:%M:%S")))
            self.attendance_table.setItem(row, 2, QTableWidgetItem(log.method_used))
            self.attendance_table.setItem(row, 3, QTableWidgetItem(log.device_id if log.device_id else "N/A"))
    
    def view_qr(self):
        # Generate QR code if it doesn't exist
        if not self.employee.qr_code:
            try:
                qr_bytes = self.db.generate_qr_code(self.employee_id)
                self.employee = self.db.get_employee(self.employee_id)  # Refresh data
                self.qr_status.setText("Généré")
                self.qr_status.setStyleSheet("color: green;")
                self.view_qr_btn.setEnabled(True)
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Échec de génération du code QR : {str(e)}")
                return
            
        # Regenerate QR code image from employee_id
        import qrcode
        from io import BytesIO
        
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(self.employee.employee_id)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Convert to QPixmap
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        qr_bytes = buffer.getvalue()
        
        qr_image = QImage()
        qr_image.loadFromData(qr_bytes)
        qr_pixmap = QPixmap.fromImage(qr_image)
        
        # Create a dialog to show the QR code
        qr_dialog = QDialog(self)
        qr_dialog.setWindowTitle(f"Code QR - {self.employee.name}")
        qr_dialog.setFixedSize(350, 400)
        
        layout = QVBoxLayout()
        
        qr_label = QLabel()
        qr_label.setPixmap(qr_pixmap.scaled(250, 250, Qt.KeepAspectRatio))
        layout.addWidget(qr_label, alignment=Qt.AlignCenter)
        
        # Employee info
        info_label = QLabel(f"Employee: {self.employee.name}\nID: {self.employee.employee_id}")
        info_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(info_label)
        
        # Save button
        save_btn = QPushButton("Save QR Image")
        save_btn.clicked.connect(lambda: self.save_qr_image(qr_pixmap))
        layout.addWidget(save_btn)
        
        qr_dialog.setLayout(layout)
        qr_dialog.exec_()
    
    def save_qr_image(self, qr_pixmap):
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Enregistrer le code QR",
            f"QR_{self.employee.employee_id}.png",
            "PNG Files (*.png)"
        )
        
        if filename:
            qr_pixmap.save(filename, "PNG")
            QMessageBox.information(self, "Succès", f"Code QR enregistré dans {filename}")
    
    def set_face(self):
        """Capture and set face image for employee using camera."""
        from utils.face_recognition import FaceRecognition
        
        face_rec = FaceRecognition()
        
        # Test camera availability
        if not face_rec.test_camera():
            QMessageBox.critical(self, "Erreur", "Aucune caméra détectée.\n\nVeuillez vous assurer qu'une caméra est connectée.")
            return
        
        # Show instructions
        QMessageBox.information(self, "Enregistrement facial", 
                               "📷 Enregistrement facial\n\n"
                               "Instructions :\n"
                               "1. Positionnez votre visage face à la caméra\n"
                               "2. Assurez-vous d'être seul dans le cadre\n"
                               "3. Regardez directement la caméra\n"
                               "4. Attendez que la qualité soit excellente (>80%)\n"
                               "5. Appuyez sur ESPACE pour capturer\n"
                               "6. Appuyez sur Q pour annuler\n\n"
                               "Conseil : Bon éclairage = meilleure reconnaissance\n"
                               "L'ancien profil facial sera remplacé.",
                               QMessageBox.Ok)
        
        # Capture face with quality validation
        result = face_rec.capture_face_for_enrollment()
        
        if result is None:
            QMessageBox.warning(self, "Erreur", "Échec de la capture faciale.")
            return
            
        face_embedding, message = result
        
        if face_embedding is None:
            QMessageBox.warning(self, "Capture annulée", f"La capture faciale a été annulée.\n\n{message}")
            return
        
        try:
            # Convert embedding numpy array to bytes for storage
            embedding_bytes = face_embedding.tobytes()
            
            # Store in database
            old_face_existed = self.employee.face_embedding is not None
            self.db.update_employee_face(self.employee_id, embedding_bytes)
            
            # Log the action
            action_type = "redéfini" if old_face_existed else "défini"
            print(f"Face {action_type} for employee {self.employee_id} - {self.employee.name}")
            
            # Refresh employee data
            self.employee = self.db.get_employee(self.employee_id)
            self.face_status.setText("Défini")
            self.face_status.setStyleSheet("color: green;")
            
            QMessageBox.information(self, "Succès", 
                                   f"✓ Visage {action_type} avec succès\n\n"
                                   f"Employé : {self.employee.name}\n"
                                   f"Le profil biométrique a été enregistré.")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Échec de l'enregistrement du visage : {str(e)}")
    
    def save_changes(self):
        new_name = self.name_edit.text().strip()
        if not new_name:
            QMessageBox.warning(self, "Erreur", "Le nom ne peut pas être vide.")
            return
            
        if new_name != self.employee.name:
            success = self.db.update_employee_name(self.employee_id, new_name)
            if success:
                QMessageBox.information(self, "Succès", "Nom de l'employé mis à jour avec succès.")
                self.employee.name = new_name
                self.setWindowTitle(f"Profil employé - {new_name}")
            else:
                QMessageBox.critical(self, "Erreur", "Échec de mise à jour du nom de l'employé.")
    
    def toggle_status(self):
        new_status = not self.employee.enabled
        self.db.update_employee_status(self.employee_id, new_status)
        self.employee.enabled = new_status
        self.status_label.setText("Activé" if new_status else "Désactivé")
        self.status_label.setStyleSheet("color: green;" if new_status else "color: red;")
        self.toggle_status_btn.setText("Désactiver" if new_status else "Activer")
        QMessageBox.information(self, "Succès", f"Employé {'activé' if new_status else 'désactivé'}.")


class AdminInterface(QWidget):
    def __init__(self, db, public):
        super().__init__()
        self.db = db
        self.public = public
        self.setWindowTitle("SLAT - Panneau d'administration")
        self.setMinimumSize(1300, 700)
        self.resize(1300, 700)  # Initial size, but resizable

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Tabs
        self.tabs = QTabWidget()
        self.layout.addWidget(self.tabs)

        # Employee Management Tab
        self.employee_tab = QWidget()
        self.setup_employee_tab()
        self.tabs.addTab(self.employee_tab, "Gestion employés")

        # Settings Tab
        self.settings_tab = QWidget()
        self.setup_settings_tab()
        self.tabs.addTab(self.settings_tab, "Paramètres")

        # Logs Tab
        self.logs_tab = QWidget()
        self.setup_logs_tab()
        self.tabs.addTab(self.logs_tab, "Journaux présence")

        # Export Filters Tab
        self.export_tab = QWidget()
        self.setup_export_tab()
        self.tabs.addTab(self.export_tab, "Filtres d'export")

        # Logout button
        logout_btn = QPushButton("Déconnexion")
        logout_btn.clicked.connect(self.logout)
        self.layout.addWidget(logout_btn)

        # Load initial data
        self.load_employees()
        self.load_logs()

    def setup_employee_tab(self):
        layout = QVBoxLayout()
        self.employee_tab.setLayout(layout)

        # Add Employee Section
        add_layout = QHBoxLayout()
        
        add_layout.addWidget(QLabel("Nom complet :"))
        self.emp_name_input = QLineEdit()
        self.emp_name_input.setPlaceholderText("Ex: Mohamed El-Amine")
        add_layout.addWidget(self.emp_name_input)

        add_btn = QPushButton("➕ Ajouter employé")
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #27AE60;
                color: white;
                padding: 8px 16px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        add_btn.clicked.connect(self.add_employee)
        add_layout.addWidget(add_btn)

        layout.addLayout(add_layout)
        
        # Info label
        info_label = QLabel("💡 L'ID employé sera généré automatiquement (Format: FP-XXXXXX)")
        info_label.setStyleSheet("color: #7F8C8D; font-size: 11px; padding: 5px;")
        layout.addWidget(info_label)

        # Employee List
        self.employee_table = QTableWidget()
        self.employee_table.setColumnCount(6)
        self.employee_table.setHorizontalHeaderLabels(["ID", "Nom", "Statut", "Code QR", "Visage", "Actions"])
        self.employee_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.employee_table.setEditTriggers(QTableWidget.NoEditTriggers)  # Make table readonly
        layout.addWidget(self.employee_table)

    def setup_settings_tab(self):
        # Create main layout
        main_layout = QVBoxLayout()
        self.settings_tab.setLayout(main_layout)
        
        # Create scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        # Create content widget
        content_widget = QWidget()
        layout = QVBoxLayout()
        content_widget.setLayout(layout)

        # Time Windows Group
        windows_group = QGroupBox("Fenêtres de pointage")
        time_layout = QVBoxLayout()
        
        time_layout.addWidget(QLabel("Fenêtre matin (arrivée) :"))

        morning_layout = QHBoxLayout()
        morning_layout.addWidget(QLabel("Début :"))
        self.morning_start = QTimeEdit()
        self.morning_start.setTime(QTime.fromString(self.db.get_setting('morning_start'), 'hh:mm'))
        morning_layout.addWidget(self.morning_start)

        morning_layout.addWidget(QLabel("Fin :"))
        self.morning_end = QTimeEdit()
        self.morning_end.setTime(QTime.fromString(self.db.get_setting('morning_end'), 'hh:mm'))
        morning_layout.addWidget(self.morning_end)

        time_layout.addLayout(morning_layout)

        time_layout.addWidget(QLabel("Fenêtre après-midi (départ) :"))

        afternoon_layout = QHBoxLayout()
        afternoon_layout.addWidget(QLabel("Début :"))
        self.afternoon_start = QTimeEdit()
        self.afternoon_start.setTime(QTime.fromString(self.db.get_setting('afternoon_start'), 'hh:mm'))
        afternoon_layout.addWidget(self.afternoon_start)

        afternoon_layout.addWidget(QLabel("Fin :"))
        self.afternoon_end = QTimeEdit()
        self.afternoon_end.setTime(QTime.fromString(self.db.get_setting('afternoon_end'), 'hh:mm'))
        afternoon_layout.addWidget(self.afternoon_end)

        time_layout.addLayout(afternoon_layout)
        windows_group.setLayout(time_layout)
        layout.addWidget(windows_group)

        # Official Work Hours Group
        official_group = QGroupBox("Horaires officiels de travail")
        official_layout = QVBoxLayout()
        
        official_desc = QLabel("Ces horaires sont utilisés pour calculer les retards et départs anticipés.")
        official_desc.setWordWrap(True)
        official_desc.setStyleSheet("color: #7F8C8D; font-size: 11px; padding: 5px;")
        official_layout.addWidget(official_desc)
        
        official_times_layout = QHBoxLayout()
        official_times_layout.addWidget(QLabel("Heure d'arrivée officielle :"))
        self.official_start_time = QTimeEdit()
        self.official_start_time.setTime(QTime.fromString(self.db.get_setting('official_start_time'), 'hh:mm'))
        official_times_layout.addWidget(self.official_start_time)
        
        official_times_layout.addWidget(QLabel("Heure de départ officielle :"))
        self.official_end_time = QTimeEdit()
        self.official_end_time.setTime(QTime.fromString(self.db.get_setting('official_end_time'), 'hh:mm'))
        official_times_layout.addWidget(self.official_end_time)
        
        official_layout.addLayout(official_times_layout)
        official_group.setLayout(official_layout)
        layout.addWidget(official_group)

        # Terminal Mode Selection
        mode_layout = QVBoxLayout()
        mode_layout.addWidget(QLabel("Mode du terminal de présence :"))
        
        mode_desc = QLabel("Le mode sélectionné détermine comment le terminal fonctionne.\n"
                          "La caméra reste active et les employés n'ont qu'à se présenter.")
        mode_desc.setWordWrap(True)
        mode_desc.setStyleSheet("color: #7F8C8D; font-size: 11px; padding: 5px;")
        mode_layout.addWidget(mode_desc)
        
        self.mode_combo = QComboBox()
        self.mode_combo.addItem("🔲 Scan QR Code", "qr")
        self.mode_combo.addItem("👤 Reconnaissance Faciale", "face")
        self.mode_combo.addItem("🔢 Saisie Carte ID", "card")
        
        current_mode = self.db.get_setting('attendance_mode')
        if current_mode == 'qr':
            self.mode_combo.setCurrentIndex(0)
        elif current_mode == 'face':
            self.mode_combo.setCurrentIndex(1)
        else:
            self.mode_combo.setCurrentIndex(2)
        
        mode_layout.addWidget(self.mode_combo)
        layout.addLayout(mode_layout)
        
        # Camera Test Section
        camera_layout = QVBoxLayout()
        camera_layout.addWidget(QLabel("Test de caméra :"))
        
        camera_desc = QLabel("Vérifiez que la caméra fonctionne correctement avant de sélectionner\n"
                           "les modes QR ou Reconnaissance Faciale.")
        camera_desc.setWordWrap(True)
        camera_desc.setStyleSheet("color: #7F8C8D; font-size: 11px; padding: 5px;")
        camera_layout.addWidget(camera_desc)
        
        test_camera_btn = QPushButton("📷 Tester la caméra")
        test_camera_btn.clicked.connect(self.test_camera)
        camera_layout.addWidget(test_camera_btn)
        
        layout.addLayout(camera_layout)
        
        # Identification Methods
        methods_layout = QVBoxLayout()
        methods_layout.addWidget(QLabel("Méthodes d'identification globales (disponibles pour tous les employés) :"))

        self.card_enabled = QCheckBox("Entrée par carte d'employé / ID")
        self.card_enabled.setChecked(self.db.get_setting('card_enabled') == '1')
        methods_layout.addWidget(self.card_enabled)

        self.qr_enabled = QCheckBox("Numérisation de code QR")
        self.qr_enabled.setChecked(self.db.get_setting('qr_enabled') == '1')
        methods_layout.addWidget(self.qr_enabled)

        self.face_enabled = QCheckBox("Reconnaissance faciale")
        self.face_enabled.setChecked(self.db.get_setting('face_enabled') == '1')
        methods_layout.addWidget(self.face_enabled)

        layout.addLayout(methods_layout)
        
        # Security Section
        security_group = QGroupBox("🔒 Sécurité")
        security_layout = QVBoxLayout()
        
        security_desc = QLabel("Mot de passe pour accéder à l'interface administrateur")
        security_desc.setStyleSheet("color: #7F8C8D; font-size: 11px; padding: 5px;")
        security_layout.addWidget(security_desc)
        
        change_pwd_btn = QPushButton("🔑 Changer le mot de passe administrateur")
        change_pwd_btn.clicked.connect(self.change_admin_password)
        security_layout.addWidget(change_pwd_btn)
        
        security_info = QLabel("⚠️ Après 3 tentatives échouées, une photo de sécurité sera prise")
        security_info.setStyleSheet("color: #E67E22; font-size: 11px; padding: 5px; font-weight: bold;")
        security_layout.addWidget(security_info)
        
        security_group.setLayout(security_layout)
        layout.addWidget(security_group)

        save_btn = QPushButton("Enregistrer les paramètres")
        save_btn.clicked.connect(self.save_settings)
        layout.addWidget(save_btn)
        
        # Set content widget to scroll area
        scroll.setWidget(content_widget)
        main_layout.addWidget(scroll)

    def setup_logs_tab(self):
        layout = QVBoxLayout()
        self.logs_tab.setLayout(layout)

        # Logs Table
        self.logs_table = QTableWidget()
        self.logs_table.setColumnCount(6)
        self.logs_table.setHorizontalHeaderLabels(["ID Employé", "Nom", "Type", "Heure", "Méthode", "Terminal"])
        self.logs_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.logs_table.setEditTriggers(QTableWidget.NoEditTriggers)  # Make table readonly
        layout.addWidget(self.logs_table)

        # Refresh and Export buttons
        button_layout = QHBoxLayout()
        refresh_btn = QPushButton("Actualiser les journaux")
        refresh_btn.clicked.connect(self.load_logs)
        button_layout.addWidget(refresh_btn)

        export_btn = QPushButton("Exporter les journaux en CSV")
        export_btn.clicked.connect(self.export_logs)
        button_layout.addWidget(export_btn)

        layout.addLayout(button_layout)

    def setup_export_tab(self):
        """Setup the export tab with preview-first design"""
        layout = QVBoxLayout()
        self.export_tab.setLayout(layout)
        
        # Main horizontal splitter (Filter Panel | Preview + Export)
        main_splitter = QSplitter(Qt.Horizontal)
        
        # === ZONE A: LEFT FILTER PANEL ===
        filter_panel = self.create_filter_panel()
        main_splitter.addWidget(filter_panel)
        
        # === ZONE B+C: RIGHT SIDE (Presets + Preview + Export) ===
        right_widget = QWidget()
        right_layout = QVBoxLayout()
        right_widget.setLayout(right_layout)
        
        # Store active preset for visual feedback
        self.active_preset = None
        
        # Zone B: Preset Buttons
        preset_group = self.create_preset_buttons()
        right_layout.addWidget(preset_group)
        
        # Zone C: Preview Table
        preview_group = self.create_preview_table()
        right_layout.addWidget(preview_group, stretch=1)
        
        # Export Actions (bottom)
        export_actions = self.create_export_actions()
        right_layout.addWidget(export_actions)
        
        # PDF Daily Attendance Sheet Export
        pdf_export_section = self.create_pdf_export_section()
        right_layout.addWidget(pdf_export_section)
        
        main_splitter.addWidget(right_widget)
        
        # Set initial splitter sizes (30% filter, 70% content)
        main_splitter.setSizes([300, 700])
        
        layout.addWidget(main_splitter)
        
        # Load initial data (current month)
        from PyQt5.QtCore import QTimer
        QTimer.singleShot(100, self.set_filter_this_month)
    
    def create_filter_panel(self):
        """Create collapsible filter panel (Zone A)"""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMaximumWidth(350)
        
        panel = QWidget()
        layout = QVBoxLayout()
        panel.setLayout(layout)
        
        # Title
        title = QLabel("🔍 Filtres")
        title.setStyleSheet("font-size: 16px; font-weight: bold; padding: 10px;")
        layout.addWidget(title)
        
        # === PÉRIODE ===
        from PyQt5.QtWidgets import QDateEdit, QListWidget, QRadioButton, QButtonGroup
        from PyQt5.QtCore import QDate
        
        period_group = QGroupBox("📅 Période")
        period_layout = QVBoxLayout()
        
        # Date inputs
        date_form = QFormLayout()
        self.filter_start_date = QDateEdit()
        self.filter_start_date.setCalendarPopup(True)
        self.filter_start_date.setDate(QDate.currentDate().addDays(-30))
        date_form.addRow("Date de début :", self.filter_start_date)
        
        self.filter_end_date = QDateEdit()
        self.filter_end_date.setCalendarPopup(True)
        self.filter_end_date.setDate(QDate.currentDate())
        date_form.addRow("Date de fin :", self.filter_end_date)
        period_layout.addLayout(date_form)
        
        # Quick shortcuts
        shortcuts_label = QLabel("Raccourcis :")
        shortcuts_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
        period_layout.addWidget(shortcuts_label)
        
        shortcut_layout = QHBoxLayout()
        btn_this_month = QPushButton("Ce mois")
        btn_this_month.clicked.connect(self.set_filter_this_month)
        shortcut_layout.addWidget(btn_this_month)
        
        btn_this_week = QPushButton("Cette semaine")
        btn_this_week.clicked.connect(self.set_filter_this_week)
        shortcut_layout.addWidget(btn_this_week)
        period_layout.addLayout(shortcut_layout)
        
        period_group.setLayout(period_layout)
        layout.addWidget(period_group)
        
        # === EMPLOYÉ ===
        employee_group = QGroupBox("👤 Employé")
        employee_layout = QVBoxLayout()
        
        self.filter_employee_combo = QComboBox()
        self.filter_employee_combo.addItem("Tous les employés", None)
        employees = self.db.get_all_employees()
        for emp in employees:
            self.filter_employee_combo.addItem(f"{emp.name} ({emp.employee_id})", emp.employee_id)
        employee_layout.addWidget(self.filter_employee_combo)
        
        employee_group.setLayout(employee_layout)
        layout.addWidget(employee_group)
        
        # === TYPE DE SITUATION ===
        situation_group = QGroupBox("📊 Type de situation")
        situation_layout = QVBoxLayout()
        
        self.filter_retards = QCheckBox("Retards")
        situation_layout.addWidget(self.filter_retards)
        
        self.filter_departs = QCheckBox("Départs anticipés")
        situation_layout.addWidget(self.filter_departs)
        
        self.filter_absences = QCheckBox("Absences")
        situation_layout.addWidget(self.filter_absences)
        
        self.filter_heures_sup = QCheckBox("Heures supplémentaires")
        situation_layout.addWidget(self.filter_heures_sup)
        
        self.filter_incomplet = QCheckBox("Journées incomplètes")
        situation_layout.addWidget(self.filter_incomplet)
        
        situation_group.setLayout(situation_layout)
        layout.addWidget(situation_group)
        
        # === RÈGLES ===
        rules_group = QGroupBox("⚙️ Règles")
        rules_layout = QFormLayout()
        
        from PyQt5.QtWidgets import QSpinBox
        self.filter_late_threshold = QSpinBox()
        self.filter_late_threshold.setRange(1, 120)
        self.filter_late_threshold.setValue(5)
        self.filter_late_threshold.setSuffix(" min")
        rules_layout.addRow("Retard >", self.filter_late_threshold)
        
        self.filter_early_threshold = QSpinBox()
        self.filter_early_threshold.setRange(1, 120)
        self.filter_early_threshold.setValue(10)
        self.filter_early_threshold.setSuffix(" min")
        rules_layout.addRow("Départ anticipé >", self.filter_early_threshold)
        
        self.filter_overtime_threshold = QSpinBox()
        self.filter_overtime_threshold.setRange(1, 240)
        self.filter_overtime_threshold.setValue(30)
        self.filter_overtime_threshold.setSuffix(" min")
        rules_layout.addRow("Heures sup >", self.filter_overtime_threshold)
        
        rules_group.setLayout(rules_layout)
        layout.addWidget(rules_group)
        
        # === MÉTHODE ===
        method_group = QGroupBox("🔐 Méthode")
        method_layout = QVBoxLayout()
        
        self.filter_method_face = QCheckBox("Face")
        method_layout.addWidget(self.filter_method_face)
        
        self.filter_method_card = QCheckBox("Carte")
        method_layout.addWidget(self.filter_method_card)
        
        self.filter_method_manual = QCheckBox("Manuel")
        method_layout.addWidget(self.filter_method_manual)
        
        method_group.setLayout(method_layout)
        layout.addWidget(method_group)
        
        # === ACTIONS ===
        action_layout = QVBoxLayout()
        action_layout.addSpacing(10)
        
        apply_btn = QPushButton("✓ Appliquer filtres")
        apply_btn.setStyleSheet("""
            QPushButton {
                background-color: #27AE60;
                color: white;
                padding: 10px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        apply_btn.clicked.connect(self.apply_filters_and_refresh)
        action_layout.addWidget(apply_btn)
        
        reset_btn = QPushButton("↺ Réinitialiser")
        reset_btn.clicked.connect(self.reset_and_refresh)
        action_layout.addWidget(reset_btn)
        
        layout.addLayout(action_layout)
        layout.addStretch()
        
        scroll.setWidget(panel)
        return scroll
    
    def create_preset_buttons(self):
        """Create preset buttons bar (Zone B)"""
        group = QGroupBox()
        layout = QVBoxLayout()
        
        title = QLabel("⚡ Raccourcis rapides")
        title.setStyleSheet("font-size: 14px; font-weight: bold; padding: 5px;")
        layout.addWidget(title)
        
        # Store preset buttons for later styling updates
        self.preset_buttons = {}
        
        # Preset buttons grid
        grid1 = QHBoxLayout()
        
        btn_all = QPushButton("📋 Toutes les présences")
        btn_all.setStyleSheet(self.get_preset_button_style("#3498DB"))
        btn_all.clicked.connect(lambda: self.apply_preset("all", btn_all))
        grid1.addWidget(btn_all)
        self.preset_buttons["all"] = (btn_all, "#3498DB")
        
        btn_late = QPushButton("🕐 Retards du mois")
        btn_late.setStyleSheet(self.get_preset_button_style("#E67E22"))
        btn_late.clicked.connect(lambda: self.apply_preset("late", btn_late))
        grid1.addWidget(btn_late)
        self.preset_buttons["late"] = (btn_late, "#E67E22")
        
        btn_early = QPushButton("🏃 Départs anticipés")
        btn_early.setStyleSheet(self.get_preset_button_style("#9B59B6"))
        btn_early.clicked.connect(lambda: self.apply_preset("early", btn_early))
        grid1.addWidget(btn_early)
        self.preset_buttons["early"] = (btn_early, "#9B59B6")
        
        layout.addLayout(grid1)
        
        grid2 = QHBoxLayout()
        
        btn_absence = QPushButton("❌ Absences")
        btn_absence.setStyleSheet(self.get_preset_button_style("#E74C3C"))
        btn_absence.clicked.connect(lambda: self.apply_preset("absence", btn_absence))
        grid2.addWidget(btn_absence)
        self.preset_buttons["absence"] = (btn_absence, "#E74C3C")
        
        btn_overtime = QPushButton("💼 Heures supplémentaires")
        btn_overtime.setStyleSheet(self.get_preset_button_style("#16A085"))
        btn_overtime.clicked.connect(lambda: self.apply_preset("overtime", btn_overtime))
        grid2.addWidget(btn_overtime)
        self.preset_buttons["overtime"] = (btn_overtime, "#16A085")
        
        btn_incomplete = QPushButton("⚠️ Journées incomplètes")
        btn_incomplete.setStyleSheet(self.get_preset_button_style("#F39C12"))
        btn_incomplete.clicked.connect(lambda: self.apply_preset("incomplete", btn_incomplete))
        grid2.addWidget(btn_incomplete)
        self.preset_buttons["incomplete"] = (btn_incomplete, "#F39C12")
        
        layout.addLayout(grid2)
        
        group.setLayout(layout)
        return group
    
    def get_preset_button_style(self, color, active=False):
        """Get consistent styling for preset buttons"""
        border = "3px solid #2ECC71" if active else "2px solid transparent"
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                padding: 15px;
                font-size: 12px;
                font-weight: bold;
                border-radius: 5px;
                min-width: 150px;
                border: {border};
            }}
            QPushButton:hover {{
                background-color: {color};
                opacity: 0.9;
            }}
            QPushButton:pressed {{
                background-color: {color};
                padding-top: 17px;
                padding-bottom: 13px;
            }}
        """
    
    def create_preview_table(self):
        """Create preview table (Zone C)"""
        group = QGroupBox("👁️ Aperçu des données")
        layout = QVBoxLayout()
        
        # Info label
        self.preview_info_label = QLabel("Aperçu chargé automatiquement...")
        self.preview_info_label.setStyleSheet("color: #7F8C8D; font-style: italic; padding: 5px;")
        layout.addWidget(self.preview_info_label)
        
        # Table
        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(8)
        self.preview_table.setHorizontalHeaderLabels([
            "Employé", "Date", "Entrée", "Sortie", "Heures", "Retard", "Heures sup", "Statut"
        ])
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.preview_table.verticalHeader().setVisible(False)
        
        # Set column widths
        header = self.preview_table.horizontalHeader()
        header.setStretchLastSection(False)
        for i in [0, 1, 2, 3, 4, 5, 6, 7]:
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents if i != 0 else QHeaderView.Stretch)
        
        layout.addWidget(self.preview_table)
        
        group.setLayout(layout)
        return group
    
    def create_export_actions(self):
        """Create export actions panel at bottom"""
        group = QGroupBox("📤 Exportation")
        layout = QHBoxLayout()
        
        # Export format (Excel only as per user request)
        format_label = QLabel("Format :")
        format_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(format_label)
        
        self.export_format_label = QLabel("📊 Excel (.xlsx)")
        self.export_format_label.setStyleSheet("color: #27AE60; font-weight: bold;")
        layout.addWidget(self.export_format_label)
        
        layout.addSpacing(20)
        
        # Mode (Daily summary only - "journal de présence")
        mode_label = QLabel("Mode :")
        mode_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(mode_label)
        
        self.export_mode_label = QLabel("📅 Journal de présence (résumé journalier)")
        self.export_mode_label.setStyleSheet("color: #27AE60; font-weight: bold;")
        layout.addWidget(self.export_mode_label)
        
        layout.addStretch()
        
        # Export button
        export_btn = QPushButton("📥 Exporter vers Excel")
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #27AE60;
                color: white;
                padding: 15px 30px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 5px;
                min-width: 200px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        export_btn.clicked.connect(self.export_to_excel)
        layout.addWidget(export_btn)
        
        group.setLayout(layout)
        return group
    
    def create_pdf_export_section(self):
        """Create PDF daily attendance sheet export section"""
        group = QGroupBox("📄 Fiche de Présence PDF")
        layout = QVBoxLayout()
        
        # Description
        desc_label = QLabel("Exportez la fiche de présence journalière en PDF")
        desc_label.setStyleSheet("color: #555; font-style: italic; margin-bottom: 10px;")
        layout.addWidget(desc_label)
        
        # Date selector and export button in horizontal layout
        h_layout = QHBoxLayout()
        
        date_label = QLabel("📅 Sélectionner le jour :")
        date_label.setStyleSheet("font-weight: bold;")
        h_layout.addWidget(date_label)
        
        self.pdf_export_date = QDateEdit()
        self.pdf_export_date.setCalendarPopup(True)
        from PyQt5.QtCore import QDate
        self.pdf_export_date.setDate(QDate.currentDate())
        self.pdf_export_date.setDisplayFormat("dd/MM/yyyy")
        h_layout.addWidget(self.pdf_export_date)
        
        h_layout.addSpacing(20)
        
        # Export PDF button
        export_pdf_btn = QPushButton("📥 Exporter Fiche de Présence PDF")
        export_pdf_btn.setStyleSheet("""
            QPushButton {
                background-color: #E74C3C;
                color: white;
                padding: 12px 25px;
                font-size: 13px;
                font-weight: bold;
                border-radius: 5px;
                min-width: 200px;
            }
            QPushButton:hover {
                background-color: #C0392B;
            }
        """)
        export_pdf_btn.clicked.connect(self.export_attendance_sheet_pdf)
        h_layout.addWidget(export_pdf_btn)
        
        h_layout.addStretch()
        
        layout.addLayout(h_layout)
        group.setLayout(layout)
        return group
    
    def set_filter_this_month(self):
        """Set filter to current month"""
        from PyQt5.QtCore import QDate
        today = QDate.currentDate()
        first_day = QDate(today.year(), today.month(), 1)
        last_day = QDate(today.year(), today.month(), calendar.monthrange(today.year(), today.month())[1])
        self.filter_start_date.setDate(first_day)
        self.filter_end_date.setDate(last_day)
        self.apply_filters_and_refresh()
    
    def set_filter_this_week(self):
        """Set filter to current week"""
        from PyQt5.QtCore import QDate
        today = QDate.currentDate()
        # Start of week (Monday)
        days_since_monday = (today.dayOfWeek() - 1) % 7
        start_of_week = today.addDays(-days_since_monday)
        end_of_week = start_of_week.addDays(6)
        self.filter_start_date.setDate(start_of_week)
        self.filter_end_date.setDate(end_of_week)
        self.apply_filters_and_refresh()
    
    def reset_filters(self):
        """Reset all filters to default"""
        from PyQt5.QtCore import QDate
        # Reset dates to current month
        today = QDate.currentDate()
        self.filter_start_date.setDate(QDate(today.year(), today.month(), 1))
        self.filter_end_date.setDate(today)
        
        # Reset employee
        self.filter_employee_combo.setCurrentIndex(0)
        
        # Uncheck all situation filters
        self.filter_retards.setChecked(False)
        self.filter_departs.setChecked(False)
        self.filter_absences.setChecked(False)
        self.filter_heures_sup.setChecked(False)
        self.filter_incomplet.setChecked(False)
        
        # Reset thresholds
        self.filter_late_threshold.setValue(5)
        self.filter_early_threshold.setValue(10)
        self.filter_overtime_threshold.setValue(30)
        
        # Uncheck all method filters
        self.filter_method_face.setChecked(False)
        self.filter_method_card.setChecked(False)
        self.filter_method_manual.setChecked(False)
    
    def reset_and_refresh(self):
        """Reset filters and refresh preview"""
        self.reset_filters()
        self.apply_filters_and_refresh()
    
    def apply_preset(self, preset_type, clicked_button=None):
        """Apply a preset filter configuration"""
        from PyQt5.QtCore import QDate
        from PyQt5.QtWidgets import QApplication
        
        # Show loading state
        preset_labels = {
            "all": "📋 Toutes les présences",
            "late": "🕐 Retards",
            "early": "🏃 Départs anticipés",
            "absence": "❌ Absences",
            "overtime": "💼 Heures supplémentaires",
            "incomplete": "⚠️ Journées incomplètes"
        }
        
        self.preview_info_label.setText(f"⏳ Chargement de {preset_labels.get(preset_type, 'données')}...")
        self.preview_info_label.setStyleSheet("color: #F39C12; font-weight: bold; padding: 5px; font-size: 13px;")
        QApplication.processEvents()  # Force UI update
        
        # Reset all button styles
        for preset_name, (btn, color) in self.preset_buttons.items():
            btn.setStyleSheet(self.get_preset_button_style(color, active=False))
        
        # Highlight the active button
        self.active_preset = preset_type
        if preset_type in self.preset_buttons:
            btn, color = self.preset_buttons[preset_type]
            btn.setStyleSheet(self.get_preset_button_style(color, active=True))
        
        # Reset all filters first (without refreshing)
        today = QDate.currentDate()
        self.filter_start_date.setDate(QDate(today.year(), today.month(), 1))
        last_day = QDate(today.year(), today.month(), calendar.monthrange(today.year(), today.month())[1])
        self.filter_end_date.setDate(last_day)
        
        # Reset employee
        self.filter_employee_combo.setCurrentIndex(0)
        
        # Uncheck all situation filters
        self.filter_retards.setChecked(False)
        self.filter_departs.setChecked(False)
        self.filter_absences.setChecked(False)
        self.filter_heures_sup.setChecked(False)
        self.filter_incomplet.setChecked(False)
        
        # Reset thresholds
        self.filter_late_threshold.setValue(5)
        self.filter_early_threshold.setValue(10)
        self.filter_overtime_threshold.setValue(30)
        
        # Uncheck all method filters
        self.filter_method_face.setChecked(False)
        self.filter_method_card.setChecked(False)
        self.filter_method_manual.setChecked(False)
        
        # Now apply the specific preset filter
        if preset_type == "all":
            # All presences - no additional filters
            pass
        elif preset_type == "late":
            # Late arrivals
            self.filter_retards.setChecked(True)
        elif preset_type == "early":
            # Early departures
            self.filter_departs.setChecked(True)
        elif preset_type == "absence":
            # Absences
            self.filter_absences.setChecked(True)
        elif preset_type == "overtime":
            # Overtime
            self.filter_heures_sup.setChecked(True)
        elif preset_type == "incomplete":
            # Incomplete days
            self.filter_incomplet.setChecked(True)
        
        # Now refresh with the preset applied
        self.apply_filters_and_refresh()
    
    def apply_filters_and_refresh(self):
        """Apply current filters and refresh preview table"""
        try:
            # Get date range
            start_date = self.filter_start_date.date().toPyDate()
            end_date = self.filter_end_date.date().toPyDate()
            
            # Get employee filter
            employee_id = self.filter_employee_combo.currentData()
            
            # Generate payroll summary for date range
            summaries = self.db.generate_payroll_summary(start_date, end_date)
            
            # Filter by employee if specified
            if employee_id:
                summaries = [s for s in summaries if s['employee_id'] == employee_id]
            
            # Apply situation filters
            filtered_summaries = []
            for summary in summaries:
                include = True
                
                # If any situation filter is checked, apply them
                if (self.filter_retards.isChecked() or self.filter_departs.isChecked() or 
                    self.filter_absences.isChecked() or self.filter_heures_sup.isChecked() or 
                    self.filter_incomplet.isChecked()):
                    
                    include = False
                    
                    # Use 'or' instead of 'if' to allow multiple conditions
                    if self.filter_retards.isChecked() and summary.get('late_minutes', 0) > self.filter_late_threshold.value():
                        include = True
                    
                    if self.filter_departs.isChecked() and summary.get('early_leave_minutes', 0) > self.filter_early_threshold.value():
                        include = True
                    
                    if self.filter_absences.isChecked() and (not summary.get('first_in') or not summary.get('last_out')):
                        include = True
                    
                    if self.filter_heures_sup.isChecked() and summary.get('overtime_minutes', 0) > self.filter_overtime_threshold.value():
                        include = True
                    
                    if self.filter_incomplet.isChecked():
                        # Check for incomplete days (mismatched IN/OUT counts)
                        # This would require additional data from database
                        pass
                
                if include:
                    filtered_summaries.append(summary)
            
            # Update preview table
            self.update_preview_table(filtered_summaries)
            
            # Update info label with descriptive message
            filter_desc = ""
            if self.filter_retards.isChecked():
                filter_desc = " avec retards"
            elif self.filter_departs.isChecked():
                filter_desc = " avec départs anticipés"
            elif self.filter_absences.isChecked():
                filter_desc = " avec absences"
            elif self.filter_heures_sup.isChecked():
                filter_desc = " avec heures supplémentaires"
            elif self.filter_incomplet.isChecked():
                filter_desc = " avec journées incomplètes"
            
            employee_desc = ""
            if self.filter_employee_combo.currentIndex() > 0:
                employee_desc = f" | Employé: {self.filter_employee_combo.currentText()}"
            
            self.preview_info_label.setText(
                f"✅ {len(filtered_summaries)} enregistrements{filter_desc} | "
                f"📅 {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}{employee_desc}"
            )
            self.preview_info_label.setStyleSheet("color: #27AE60; font-weight: bold; padding: 5px; font-size: 13px;")
            
        except Exception as e:
            import traceback
            error_msg = f"Erreur lors de l'actualisation :\n{str(e)}\n\n{traceback.format_exc()}"
            print(error_msg)  # Print to console for debugging
            QMessageBox.warning(self, "Erreur", f"Erreur lors de l'actualisation : {str(e)}")
    
    def update_preview_table(self, summaries):
        """Update preview table with summary data"""
        self.preview_table.setRowCount(0)
        
        for summary in summaries:
            row = self.preview_table.rowCount()
            self.preview_table.insertRow(row)
            
            # Employee name
            self.preview_table.setItem(row, 0, QTableWidgetItem(summary.get('employee_name', 'N/A')))
            
            # Date
            self.preview_table.setItem(row, 1, QTableWidgetItem(summary.get('date', 'N/A')))
            
            # First IN
            first_in = summary.get('first_in') if summary.get('first_in') else 'N/A'
            self.preview_table.setItem(row, 2, QTableWidgetItem(first_in))
            
            # Last OUT
            last_out = summary.get('last_out') if summary.get('last_out') else 'N/A'
            self.preview_table.setItem(row, 3, QTableWidgetItem(last_out))
            
            # Hours worked
            hours_worked = summary.get('hours_worked', 0)
            hours = f"{hours_worked:.2f}h" if hours_worked else "0.00h"
            self.preview_table.setItem(row, 4, QTableWidgetItem(hours))
            
            # Late minutes
            late_minutes = summary.get('late_minutes', 0)
            late_item = QTableWidgetItem(f"{late_minutes} min")
            if late_minutes > 0:
                late_item.setBackground(QBrush(QColor(231, 76, 60, 80)))  # Red
            self.preview_table.setItem(row, 5, late_item)
            
            # Overtime minutes
            overtime_minutes = summary.get('overtime_minutes', 0)
            overtime_item = QTableWidgetItem(f"{overtime_minutes} min")
            if overtime_minutes > 0:
                overtime_item.setBackground(QBrush(QColor(52, 152, 219, 80)))  # Blue
            self.preview_table.setItem(row, 6, overtime_item)
            
            # Status
            status = "✓ Complet"
            status_color = None
            early_leave = summary.get('early_leave_minutes', 0)
            
            if not summary.get('first_in') or not summary.get('last_out'):
                status = "⚫ Absent"
                status_color = QColor(149, 165, 166, 80)  # Grey
            elif late_minutes > 0:
                status = "🔴 Retard"
                status_color = QColor(231, 76, 60, 80)  # Red
            elif early_leave > 0:
                status = "🟠 Départ anticipé"
                status_color = QColor(230, 126, 34, 80)  # Orange
            elif overtime_minutes > 30:
                status = "🔵 Heures sup"
                status_color = QColor(52, 152, 219, 80)  # Blue
            
            status_item = QTableWidgetItem(status)
            if status_color:
                status_item.setBackground(QBrush(status_color))
            self.preview_table.setItem(row, 7, status_item)
    
    def export_to_excel(self):
        """Export current preview data to Excel"""
        if self.preview_table.rowCount() == 0:
            QMessageBox.warning(self, "Attention", "Aucune donnée à exporter. Veuillez appliquer des filtres d'abord.")
            return
        
        # Get filename
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter vers Excel",
            f"journal_presence_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Journal de Présence"
            
            # Headers
            headers = ["Employé", "Date", "Entrée", "Sortie", "Heures", "Retard (min)", "Heures sup (min)", "Statut"]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Add data from preview table
            for row in range(self.preview_table.rowCount()):
                row_data = []
                for col in range(self.preview_table.columnCount()):
                    item = self.preview_table.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save
            wb.save(filename)
            
            # Show success message
            QMessageBox.information(self, "Succès",
                f"✅ Fichier généré avec succès!\n\n"
                f"📁 Fichier: {filename}\n"
                f"📊 Lignes exportées: {self.preview_table.rowCount()}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Échec de l'exportation : {str(e)}")
    
    def export_attendance_sheet_pdf(self):
        """Export daily attendance sheet to PDF (Fiche de Présence)"""
        selected_date = self.pdf_export_date.date().toPyDate()
        
        # Get attendance records for the selected day
        with sqlite3.connect(self.db.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT 
                    e.name as employee_name,
                    al.timestamp,
                    al.type,
                    al.method
                FROM attendance_logs al
                JOIN employees e ON al.employee_id = e.employee_id
                WHERE DATE(al.timestamp) = ?
                AND al.status = 'ACCEPTED'
                ORDER BY e.name, al.timestamp
            ''', (selected_date.strftime('%Y-%m-%d'),))
            
            records = cursor.fetchall()
        
        if not records:
            QMessageBox.warning(self, "Attention", 
                f"Aucune donnée de présence trouvée pour le {selected_date.strftime('%d/%m/%Y')}")
            return
        
        # Group records by employee
        from collections import defaultdict
        employee_records = defaultdict(lambda: {'in': None, 'out': None})
        
        for name, timestamp, record_type, method in records:
            timestamp_dt = datetime.fromisoformat(timestamp)
            if record_type == 'IN' and employee_records[name]['in'] is None:
                employee_records[name]['in'] = (timestamp_dt.strftime('%H:%M:%S'), method)
            elif record_type == 'OUT':
                employee_records[name]['out'] = (timestamp_dt.strftime('%H:%M:%S'), method)
        
        # Get filename
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter Fiche de Présence PDF",
            f"fiche_presence_{selected_date.strftime('%Y%m%d')}.pdf",
            "PDF Files (*.pdf)"
        )
        
        if not filename:
            return
        
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib.enums import TA_CENTER
            
            # Create PDF
            doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#2C3E50'),
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            title_text = f"FICHE DE PRÉSENCE<br/>{selected_date.strftime('%d/%m/%Y')}"
            title = Paragraph(title_text, title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.5*cm))
            
            # Table data
            data = [['NOM EMPLOYÉ', 'H-D\'ENTRÉE', 'TYPE ENTRÉE', 'H-DE SORTIE', 'TYPE SORTIE']]
            
            for emp_name in sorted(employee_records.keys()):
                record = employee_records[emp_name]
                entry_time = record['in'][0] if record['in'] else '-'
                entry_type = self.format_method_name(record['in'][1]) if record['in'] else '-'
                exit_time = record['out'][0] if record['out'] else '-'
                exit_type = self.format_method_name(record['out'][1]) if record['out'] else '-'
                
                data.append([
                    emp_name,
                    entry_time,
                    entry_type,
                    exit_time,
                    exit_type
                ])
            
            # Create table
            table = Table(data, colWidths=[6*cm, 3*cm, 3*cm, 3*cm, 3*cm])
            
            # Style table
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ECF0F1')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ]))
            
            elements.append(table)
            
            # Footer
            elements.append(Spacer(1, 1*cm))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#7F8C8D'),
                alignment=TA_CENTER
            )
            footer_text = f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} - SLAT v1.0"
            footer = Paragraph(footer_text, footer_style)
            elements.append(footer)
            
            # Build PDF
            doc.build(elements)
            
            # Show success message
            QMessageBox.information(self, "Succès",
                f"✅ Fiche de présence générée avec succès!\n\n"
                f"📁 Fichier: {filename}\n"
                f"📅 Date: {selected_date.strftime('%d/%m/%Y')}\n"
                f"👥 Employés: {len(employee_records)}")
            
        except ImportError:
            QMessageBox.critical(self, "Erreur", 
                "La bibliothèque ReportLab n'est pas installée.\n\n"
                "Installez-la avec: pip install reportlab")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Échec de l'exportation PDF: {str(e)}")
    
    def format_method_name(self, method):
        """Format method name for display"""
        method_map = {
            'FACE': '👤 VISAGE',
            'QR': '📱 QR CODE',
            'CARD': '💳 CARTE',
            'MANUAL': '✍️ MANUEL'
        }
        return method_map.get(method, method)
    
    def load_employees_to_combo(self):
        """Load employees into single employee combo box"""
        self.single_employee_combo.clear()
        employees = self.db.get_all_employees()
        for emp in employees:
            self.single_employee_combo.addItem(f"{emp.name} ({emp.employee_id})", emp.employee_id)
    
    def load_employees_to_list(self):
        """Load employees into multiple selection list"""
        self.multiple_employees_list.clear()
        employees = self.db.get_all_employees()
        for emp in employees:
            self.multiple_employees_list.addItem(f"{emp.name} ({emp.employee_id})")
    
    def get_filter_date_range(self):
        """Get the date range based on selected period filter"""
        selected_period = self.period_type_group.checkedId()
        
        if selected_period == 1:  # Month
            month = self.month_combo.currentIndex() + 1
            year = int(self.year_combo.currentText())
            
            # First day of month
            start_date = datetime(year, month, 1).date()
            
            # Last day of month
            if month == 12:
                end_date = datetime(year, 12, 31).date()
            else:
                end_date = datetime(year, month + 1, 1).date()
                end_date = end_date.replace(day=end_date.day - 1) if end_date.day > 1 else end_date
            
            # Simple approach: just get last day of month
            import calendar
            last_day = calendar.monthrange(year, month)[1]
            end_date = datetime(year, month, last_day).date()
            
        elif selected_period == 2:  # Week
            start_date = self.week_start_date.date().toPyDate()
            end_date = start_date.replace(day=start_date.day + 6)
            
        else:  # Custom
            start_date = self.custom_start_date.date().toPyDate()
            end_date = self.custom_end_date.date().toPyDate()
        
        return start_date, end_date
    
    def get_filter_employees(self):
        """Get the list of employee IDs based on selected employee filter"""
        selected_type = self.employee_type_group.checkedId()
        
        if selected_type == 1:  # All employees
            return None  # None means all
        elif selected_type == 2:  # Single employee
            return [self.single_employee_combo.currentData()]
        else:  # Multiple employees
            selected_items = self.multiple_employees_list.selectedItems()
            employee_ids = []
            for item in selected_items:
                # Extract employee_id from "Name (ID)" format
                text = item.text()
                emp_id = text[text.rfind('(') + 1:text.rfind(')')]
                employee_ids.append(emp_id)
            return employee_ids if employee_ids else None
    
    def export_filtered_audit_trail(self):
        """Export audit trail with filters"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter audit trail",
            f"audit_trail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if filename:
            try:
                start_date, end_date = self.get_filter_date_range()
                employee_ids = self.get_filter_employees()
                
                # For now, use the basic export and filter in future
                self.db.export_audit_trail_csv(start_date, end_date, filename)
                
                QMessageBox.information(self, "Succès", 
                    f"Audit trail exporté avec succès!\n\n"
                    f"Période: {start_date} à {end_date}\n"
                    f"Fichier: {filename}")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Échec d'exportation : {str(e)}")
    
    def export_filtered_payroll_csv(self):
        """Export payroll summary to CSV with filters"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter données paie (CSV)",
            f"payroll_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if filename:
            try:
                start_date, end_date = self.get_filter_date_range()
                employee_ids = self.get_filter_employees()
                
                # Export payroll
                self.db.export_payroll_csv(start_date, end_date, filename)
                
                QMessageBox.information(self, "Succès",
                    f"Données paie exportées avec succès!\n\n"
                    f"Période: {start_date} à {end_date}\n"
                    f"Format: CSV\n"
                    f"Fichier: {filename}")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Échec d'exportation : {str(e)}")
    
    def export_filtered_payroll_excel(self):
        """Export payroll summary to Excel with filters"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter données paie (Excel)",
            f"payroll_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if filename:
            try:
                start_date, end_date = self.get_filter_date_range()
                employee_ids = self.get_filter_employees()
                
                # Export payroll
                self.db.export_payroll_excel(start_date, end_date, filename)
                
                QMessageBox.information(self, "Succès",
                    f"Données paie exportées avec succès!\n\n"
                    f"Période: {start_date} à {end_date}\n"
                    f"Format: Excel (XLSX)\n"
                    f"Fichier: {filename}")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Échec d'exportation : {str(e)}")
    
    def export_late_arrivals_report(self):
        """Export late arrivals report"""
        # Show filter dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Filtres pour rapport retards")
        dialog.setFixedWidth(400)
        
        layout = QVBoxLayout()
        dialog.setLayout(layout)
        
        layout.addWidget(QLabel("Sélectionnez les critères de retard :"))
        
        late_5min = QCheckBox("Retard > 5 minutes")
        layout.addWidget(late_5min)
        
        late_10min = QCheckBox("Retard > 10 minutes")
        layout.addWidget(late_10min)
        
        late_30min = QCheckBox("Retard > 30 minutes")
        layout.addWidget(late_30min)
        
        all_lates = QCheckBox("Tous les retards (>0 min)")
        all_lates.setChecked(True)
        layout.addWidget(all_lates)
        
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("Exporter")
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn = QPushButton("Annuler")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(ok_btn)
        layout.addLayout(btn_layout)
        
        if dialog.exec_() != QDialog.Accepted:
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter rapport retards",
            f"late_arrivals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if filename:
            try:
                import csv
                
                start_date, end_date = self.get_filter_date_range()
                summaries = self.db.generate_payroll_summary(start_date, end_date)
                
                # Filter based on selected criteria
                filtered = []
                for summary in summaries:
                    late_min = summary['late_minutes']
                    
                    include = False
                    if all_lates.isChecked() and late_min > 0:
                        include = True
                    elif late_5min.isChecked() and late_min > 5:
                        include = True
                    elif late_10min.isChecked() and late_min > 10:
                        include = True
                    elif late_30min.isChecked() and late_min > 30:
                        include = True
                    
                    if include:
                        filtered.append(summary)
                
                # Get official start time
                official_start = self.db.get_setting('official_start_time')
                
                # Write CSV
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Employee ID', 'Name', 'Date', 'Official Time', 'Actual Arrival', 'Late Minutes'])
                    
                    for item in filtered:
                        writer.writerow([
                            item['employee_id'],
                            item['employee_name'],
                            item['date'],
                            official_start,
                            item['first_in'],
                            item['late_minutes']
                        ])
                
                QMessageBox.information(self, "Succès",
                    f"✅ Rapport retards exporté!\n\n"
                    f"📅 Période: {start_date} à {end_date}\n"
                    f"📊 Total: {len(filtered)} enregistrements\n"
                    f"📁 Fichier: {filename}")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Échec d'exportation : {str(e)}")
    
    def export_early_departures_report(self):
        """Export early departures report"""
        # Show filter dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Filtres pour rapport sorties anticipées")
        dialog.setFixedWidth(400)
        
        layout = QVBoxLayout()
        dialog.setLayout(layout)
        
        layout.addWidget(QLabel("Sélectionnez les critères de sortie anticipée :"))
        
        early_5min = QCheckBox("Sortie anticipée > 5 minutes")
        layout.addWidget(early_5min)
        
        early_15min = QCheckBox("Sortie anticipée > 15 minutes")
        layout.addWidget(early_15min)
        
        early_30min = QCheckBox("Sortie anticipée > 30 minutes")
        layout.addWidget(early_30min)
        
        all_early = QCheckBox("Toutes les sorties anticipées (>0 min)")
        all_early.setChecked(True)
        layout.addWidget(all_early)
        
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("Exporter")
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn = QPushButton("Annuler")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(ok_btn)
        layout.addLayout(btn_layout)
        
        if dialog.exec_() != QDialog.Accepted:
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter rapport départs anticipés",
            f"early_departures_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if filename:
            try:
                import csv
                
                start_date, end_date = self.get_filter_date_range()
                summaries = self.db.generate_payroll_summary(start_date, end_date)
                
                # Filter based on selected criteria
                filtered = []
                for summary in summaries:
                    early_min = summary['early_leave_minutes']
                    
                    include = False
                    if all_early.isChecked() and early_min > 0:
                        include = True
                    elif early_5min.isChecked() and early_min > 5:
                        include = True
                    elif early_15min.isChecked() and early_min > 15:
                        include = True
                    elif early_30min.isChecked() and early_min > 30:
                        include = True
                    
                    if include:
                        filtered.append(summary)
                
                # Get official end time
                official_end = self.db.get_setting('official_end_time')
                
                # Write CSV
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Employee ID', 'Name', 'Date', 'Official Time', 'Actual Departure', 'Early Leave Minutes'])
                    
                    for item in filtered:
                        writer.writerow([
                            item['employee_id'],
                            item['employee_name'],
                            item['date'],
                            official_end,
                            item['last_out'],
                            item['early_leave_minutes']
                        ])
                
                QMessageBox.information(self, "Succès",
                    f"✅ Rapport départs anticipés exporté!\n\n"
                    f"📅 Période: {start_date} à {end_date}\n"
                    f"📊 Total: {len(filtered)} enregistrements\n"
                    f"📁 Fichier: {filename}")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Échec d'exportation : {str(e)}")
    
    def export_absences_report(self):
        """Export absences report"""
        # Show filter dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Filtres pour rapport absences")
        dialog.setFixedWidth(400)
        
        layout = QVBoxLayout()
        dialog.setLayout(layout)
        
        layout.addWidget(QLabel("Sélectionnez les types d'absence :"))
        
        absence_complete = QCheckBox("Absence complète (pas de IN/OUT)")
        layout.addWidget(absence_complete)
        
        absence_no_in = QCheckBox("Pas de pointage IN")
        layout.addWidget(absence_no_in)
        
        absence_no_out = QCheckBox("Pas de pointage OUT")
        layout.addWidget(absence_no_out)
        
        all_absences = QCheckBox("Toutes les absences")
        all_absences.setChecked(True)
        layout.addWidget(all_absences)
        
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("Exporter")
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn = QPushButton("Annuler")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(ok_btn)
        layout.addLayout(btn_layout)
        
        if dialog.exec_() != QDialog.Accepted:
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter rapport absences",
            f"absences_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if filename:
            try:
                import csv
                
                start_date, end_date = self.get_filter_date_range()
                summaries = self.db.generate_payroll_summary(start_date, end_date)
                
                # Filter based on selected criteria
                filtered = []
                for summary in summaries:
                    has_in = summary['first_in'] is not None
                    has_out = summary['last_out'] is not None
                    
                    include = False
                    if all_absences.isChecked() and (not has_in or not has_out):
                        include = True
                    elif absence_complete.isChecked() and not has_in and not has_out:
                        include = True
                    elif absence_no_in.isChecked() and not has_in:
                        include = True
                    elif absence_no_out.isChecked() and not has_out:
                        include = True
                    
                    if include:
                        filtered.append(summary)
                
                # Write CSV
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Employee ID', 'Name', 'Date', 'First IN', 'Last OUT', 'Status'])
                    
                    for item in filtered:
                        status = 'Absent complet'
                        if item['first_in'] and not item['last_out']:
                            status = 'Pas de OUT'
                        elif not item['first_in'] and item['last_out']:
                            status = 'Pas de IN'
                        
                        writer.writerow([
                            item['employee_id'],
                            item['employee_name'],
                            item['date'],
                            item['first_in'] if item['first_in'] else 'N/A',
                            item['last_out'] if item['last_out'] else 'N/A',
                            status
                        ])
                
                QMessageBox.information(self, "Succès",
                    f"✅ Rapport absences exporté!\n\n"
                    f"📅 Période: {start_date} à {end_date}\n"
                    f"📊 Total: {len(filtered)} enregistrements\n"
                    f"📁 Fichier: {filename}")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Échec d'exportation : {str(e)}")
    
    def export_incomplete_days_report(self):
        """Export incomplete days report"""
        # Show filter dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Filtres pour rapport journées incomplètes")
        dialog.setFixedWidth(400)
        
        layout = QVBoxLayout()
        dialog.setLayout(layout)
        
        layout.addWidget(QLabel("Sélectionnez les types d'anomalies :"))
        
        incomplete_in_no_out = QCheckBox("IN sans OUT")
        layout.addWidget(incomplete_in_no_out)
        
        incomplete_out_no_in = QCheckBox("OUT sans IN")
        layout.addWidget(incomplete_out_no_in)
        
        incomplete_odd_punches = QCheckBox("Nombre impair de pointages")
        layout.addWidget(incomplete_odd_punches)
        
        all_incomplete = QCheckBox("Toutes les journées incomplètes")
        all_incomplete.setChecked(True)
        layout.addWidget(all_incomplete)
        
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("Exporter")
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn = QPushButton("Annuler")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(ok_btn)
        layout.addLayout(btn_layout)
        
        if dialog.exec_() != QDialog.Accepted:
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter rapport journées incomplètes",
            f"incomplete_days_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if filename:
            try:
                import csv
                import sqlite3
                
                start_date, end_date = self.get_filter_date_range()
                
                # Get raw attendance data to check for incomplete days
                filtered = []
                employees = self.db.get_all_employees()
                
                with sqlite3.connect(self.db.db_path) as conn:
                    cursor = conn.cursor()
                    
                    for emp in employees:
                        # Get all punches for this employee in date range
                        cursor.execute('''
                            SELECT DATE(timestamp), type, TIME(timestamp), COUNT(*)
                            FROM attendance_logs
                            WHERE employee_id = ?
                            AND DATE(timestamp) BETWEEN ? AND ?
                            AND status = 'ACCEPTED'
                            GROUP BY DATE(timestamp), type
                            ORDER BY timestamp
                        ''', (emp.employee_id, start_date, end_date))
                        
                        daily_data = {}
                        for row in cursor.fetchall():
                            date = row[0]
                            punch_type = row[1]
                            
                            if date not in daily_data:
                                daily_data[date] = {'in_count': 0, 'out_count': 0}
                            
                            if punch_type == 'IN':
                                daily_data[date]['in_count'] += 1
                            else:
                                daily_data[date]['out_count'] += 1
                        
                        # Check for incomplete days
                        for date, counts in daily_data.items():
                            in_count = counts['in_count']
                            out_count = counts['out_count']
                            
                            include = False
                            issue = ""
                            
                            if all_incomplete.isChecked() and in_count != out_count:
                                include = True
                                if in_count > out_count:
                                    issue = "IN sans OUT"
                                else:
                                    issue = "OUT sans IN"
                            elif incomplete_in_no_out.isChecked() and in_count > 0 and out_count == 0:
                                include = True
                                issue = "IN sans OUT"
                            elif incomplete_out_no_in.isChecked() and out_count > 0 and in_count == 0:
                                include = True
                                issue = "OUT sans IN"
                            elif incomplete_odd_punches.isChecked() and (in_count + out_count) % 2 != 0:
                                include = True
                                issue = "Nombre impair de pointages"
                            
                            if include:
                                filtered.append({
                                    'employee_id': emp.employee_id,
                                    'name': emp.name,
                                    'date': date,
                                    'in_count': in_count,
                                    'out_count': out_count,
                                    'issue': issue
                                })
                
                # Write CSV
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Employee ID', 'Name', 'Date', 'IN Count', 'OUT Count', 'Issue'])
                    
                    for item in filtered:
                        writer.writerow([
                            item['employee_id'],
                            item['name'],
                            item['date'],
                            item['in_count'],
                            item['out_count'],
                            item['issue']
                        ])
                
                QMessageBox.information(self, "Succès",
                    f"✅ Rapport journées incomplètes exporté!\n\n"
                    f"📅 Période: {start_date} à {end_date}\n"
                    f"📊 Total: {len(filtered)} enregistrements\n"
                    f"📁 Fichier: {filename}")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Échec d'exportation : {str(e)}")
                
                QMessageBox.information(self, "Succès",
                    f"Rapport journées incomplètes exporté!\n\n"
                    f"Période: {start_date} à {end_date}\n"
                    f"Total: {len(filtered)} enregistrements\n"
                    f"Fichier: {filename}")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Échec d'exportation : {str(e)}")

    def add_employee(self):
        name = self.emp_name_input.text().strip()
        
        if not name:
            QMessageBox.warning(self, "Erreur", "Veuillez saisir le nom de l'employé.")
            return
        
        # Generate professional employee ID
        emp_id = self.generate_employee_id()
        
        # Add to database
        success = self.db.add_employee(emp_id, name)
        if success:
            QMessageBox.information(self, "Succès", 
                                   f"✓ Employé ajouté avec succès!\n\n"
                                   f"Nom: {name}\n"
                                   f"ID: {emp_id}\n\n"
                                   f"Vous pouvez maintenant configurer le QR code et le visage.")
            self.emp_name_input.clear()
            self.load_employees()
        else:
            QMessageBox.warning(self, "Erreur", "Échec de l'ajout de l'employé.")

    def generate_employee_id(self):
        """Generate a unique professional employee ID in format FP-XXXXXX"""
        import random
        
        while True:
            # Generate 6-digit number
            number = random.randint(100000, 999999)
            emp_id = f"FP-{number}"
            
            # Check if ID already exists
            if not self.db.get_employee(emp_id):
                return emp_id

    def view_profile(self, employee_id):
        """Open employee profile dialog"""
        dialog = EmployeeProfileDialog(self.db, employee_id)
        dialog.exec_()
        # Refresh the table after dialog closes
        self.load_employees()

    def load_employees(self):
        """Load all employees into the table"""
        employees = self.db.get_all_employees()
        self.employee_table.setRowCount(len(employees))
        
        for row, emp in enumerate(employees):
            self.employee_table.setItem(row, 0, QTableWidgetItem(emp.employee_id))
            self.employee_table.setItem(row, 1, QTableWidgetItem(emp.name))
            self.employee_table.setItem(row, 2, QTableWidgetItem("Enabled" if emp.enabled else "Disabled"))
            self.employee_table.setItem(row, 3, QTableWidgetItem("✓" if emp.qr_code else "✗"))
            self.employee_table.setItem(row, 4, QTableWidgetItem("✓" if emp.face_embedding else "✗"))
            
            # Action buttons
            action_widget = QWidget()
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(2, 2, 2, 2)
            
            # Profile button
            profile_btn = QPushButton("Profile")
            profile_btn.setToolTip("Voir/Modifier le profil employé")
            profile_btn.clicked.connect(lambda checked, eid=emp.employee_id: self.view_profile(eid))
            action_layout.addWidget(profile_btn)
            
            action_widget.setLayout(action_layout)
            self.employee_table.setCellWidget(row, 5, action_widget)
    
    def generate_qr(self, employee_id):
        """Generate and save QR code for employee"""
        try:
            qr_bytes = self.db.generate_qr_code(employee_id)
            
            # Save QR code image
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Save QR Code",
                f"QR_{employee_id}.png",
                "PNG Files (*.png)"
            )
            
            if filename:
                with open(filename, 'wb') as f:
                    f.write(qr_bytes)
                QMessageBox.information(self, "Succès", f"Code QR généré et enregistré dans {filename}")
                self.load_employees()
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Échec de génération du code QR : {str(e)}")
    
    def set_face(self, employee_id):
        """Set face image for employee using camera"""
        from utils.face_recognition import FaceRecognition
        
        face_rec = FaceRecognition()
        
        # Get employee info
        emp = self.db.get_employee(employee_id)
        if not emp:
            QMessageBox.critical(self, "Erreur", "Employé introuvable.")
            return
        
        # Test camera
        if not face_rec.test_camera():
            QMessageBox.critical(self, "Erreur", "Aucune caméra détectée.")
            return
        
        # Show instructions
        QMessageBox.information(self, "Enregistrement facial", 
                               f"📷 Enregistrement facial pour {emp.name}\n\n"
                               "Instructions :\n"
                               "1. Positionnez-vous face à la caméra\n"
                               "2. Bon éclairage requis\n"
                               "3. Attendez qualité >80%\n"
                               "4. Appuyez sur ESPACE pour capturer\n"
                               "5. Appuyez sur Q pour annuler",
                               QMessageBox.Ok)
        
        # Capture face
        result = face_rec.capture_face_for_enrollment()
        
        if result is None:
            QMessageBox.warning(self, "Erreur", "Échec de la capture faciale.")
            return
            
        face_data, message = result
        
        if face_data is None:
            QMessageBox.warning(self, "Capture annulée", f"Capture annulée.\n\n{message}")
            return
        
        try:
            face_bytes = face_data.tobytes()
            self.db.update_employee_face(employee_id, face_bytes)
            QMessageBox.information(self, "Succès", f"Image faciale définie pour {emp.name}")
            self.load_employees()
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Échec de définition de l'image faciale : {str(e)}")

    def toggle_employee(self, employee_id, new_status):
        """Enable or disable an employee"""
        self.db.update_employee_status(employee_id, new_status)
        self.load_employees()
        QMessageBox.information(self, "Succès", f"Employé {employee_id} {'activé' if new_status else 'désactivé'}.")

    def load_logs(self):
        """Load attendance logs into the table"""
        logs = self.db.get_all_logs(limit=100)
        self.logs_table.setRowCount(len(logs))
        
        for row, log in enumerate(logs):
            # New schema: id, record_id, employee_id, terminal_id, timestamp, type, method, 
            #            confidence, status, operator_id, correction_reason, replaces_record_id, 
            #            photo_path, integrity_hash, created_at, modified_at
            emp = self.db.get_employee(log[2])
            emp_name = emp.name if emp else "Unknown"
            
            confidence_str = f"{log[7]:.1f}%" if log[7] is not None else "N/A"
            
            self.logs_table.setItem(row, 0, QTableWidgetItem(log[2]))  # employee_id
            self.logs_table.setItem(row, 1, QTableWidgetItem(emp_name))  # name
            self.logs_table.setItem(row, 2, QTableWidgetItem(log[5]))  # type (IN/OUT)
            self.logs_table.setItem(row, 3, QTableWidgetItem(str(log[4])))  # timestamp
            self.logs_table.setItem(row, 4, QTableWidgetItem(log[6]))  # method
            self.logs_table.setItem(row, 5, QTableWidgetItem(log[3] if log[3] else "N/A"))  # terminal_id

    def change_admin_password(self):
        """Change the admin password with verification"""
        import hashlib
        from PyQt5.QtWidgets import QInputDialog
        
        # Get the stored password hash (default is SHA-256 hash of "admin")
        stored_hash = self.db.get_setting('admin_password')
        if not stored_hash:
            stored_hash = hashlib.sha256("admin".encode()).hexdigest()
            self.db.update_setting('admin_password', stored_hash)
        
        # Step 1: Verify current password
        current_password, ok = QInputDialog.getText(
            self, 
            "Mot de passe actuel", 
            "Entrez le mot de passe administrateur actuel:",
            QLineEdit.Password
        )
        
        if not ok:
            return
        
        # Hash the entered current password and verify
        current_hash = hashlib.sha256(current_password.encode()).hexdigest()
        if current_hash != stored_hash:
            QMessageBox.warning(self, "Erreur", "❌ Mot de passe actuel incorrect!")
            return
        
        # Step 2: Get new password
        new_password, ok = QInputDialog.getText(
            self, 
            "Nouveau mot de passe", 
            "Entrez le nouveau mot de passe administrateur:\n(Minimum 4 caractères)",
            QLineEdit.Password
        )
        
        if not ok:
            return
        
        # Validate password length
        if len(new_password) < 4:
            QMessageBox.warning(self, "Erreur", "❌ Le mot de passe doit contenir au moins 4 caractères!")
            return
        
        # Step 3: Confirm new password
        confirm_password, ok = QInputDialog.getText(
            self, 
            "Confirmer le mot de passe", 
            "Confirmez le nouveau mot de passe:",
            QLineEdit.Password
        )
        
        if not ok:
            return
        
        # Verify passwords match
        if new_password != confirm_password:
            QMessageBox.warning(self, "Erreur", "❌ Les mots de passe ne correspondent pas!")
            return
        
        # Hash and save the new password
        new_hash = hashlib.sha256(new_password.encode()).hexdigest()
        self.db.update_setting('admin_password', new_hash)
        
        QMessageBox.information(self, "Succès", "✅ Mot de passe administrateur changé avec succès!\n\n"
                                              "Le nouveau mot de passe sera requis lors de la prochaine connexion.")

    def save_settings(self):
        # Save time windows
        self.db.update_setting('morning_start', self.morning_start.time().toString('HH:mm'))
        self.db.update_setting('morning_end', self.morning_end.time().toString('HH:mm'))
        self.db.update_setting('afternoon_start', self.afternoon_start.time().toString('HH:mm'))
        self.db.update_setting('afternoon_end', self.afternoon_end.time().toString('HH:mm'))
        
        # Save official work hours
        self.db.update_setting('official_start_time', self.official_start_time.time().toString('HH:mm'))
        self.db.update_setting('official_end_time', self.official_end_time.time().toString('HH:mm'))

        # Save attendance mode
        mode_data = self.mode_combo.currentData()
        self.db.update_setting('attendance_mode', mode_data)

        # Save identification method settings
        self.db.update_setting('card_enabled', '1' if self.card_enabled.isChecked() else '0')
        self.db.update_setting('qr_enabled', '1' if self.qr_enabled.isChecked() else '0')
        self.db.update_setting('face_enabled', '1' if self.face_enabled.isChecked() else '0')

        QMessageBox.information(self, "Succès", "Paramètres enregistrés avec succès.\n\n⚠ Redémarrez le terminal de présence pour appliquer le nouveau mode.")

    def test_camera(self):
        """Test camera availability with robust initialization"""
        import cv2
        
        # Try different camera indices
        for camera_index in [0, 1, 2]:
            try:
                cap = cv2.VideoCapture(camera_index, cv2.CAP_DSHOW)
                
                if cap.isOpened():
                    # Test if we can actually read frames
                    ret, test_frame = cap.read()
                    if ret and test_frame is not None:
                        cap.release()
                        QMessageBox.information(self, "Test caméra réussi", 
                                              f"✅ Caméra détectée sur l'index {camera_index}\n\n"
                                              f"Résolution: {test_frame.shape[1]}x{test_frame.shape[0]}\n\n"
                                              f"Le mode QR et Reconnaissance Faciale devraient fonctionner.")
                        return
                    else:
                        cap.release()
                        continue
                else:
                    continue
                    
            except Exception as e:
                print(f"Failed to initialize camera {camera_index}: {e}")
                continue
        
        # If all cameras failed
        QMessageBox.warning(self, "Test caméra échoué", 
                          "❌ Aucune caméra détectée\n\n"
                          "Vérifiez que :\n"
                          "• La caméra est connectée\n"
                          "• Les pilotes sont installés\n"
                          "• L'application a les permissions caméra\n\n"
                          "Seul le mode Carte ID sera disponible.")

    def export_logs(self):
        """Export logs to CSV file"""
        filename, _ = QFileDialog.getSaveFileName(
            self, 
            "Exporter les journaux", 
            f"attendance_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if filename:
            try:
                logs = self.db.get_all_logs()
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['Record ID', 'Employee ID', 'Name', 'Terminal ID', 'Timestamp', 
                                   'Type', 'Method', 'Confidence', 'Status', 'Photo Path', 'Integrity Hash'])
                    
                    for log in logs:
                        # New schema: id, record_id, employee_id, terminal_id, timestamp, type, method, 
                        #            confidence, status, operator_id, correction_reason, replaces_record_id, 
                        #            photo_path, integrity_hash, created_at, modified_at
                        emp = self.db.get_employee(log[2])
                        emp_name = emp.name if emp else "Unknown"
                        writer.writerow([
                            log[1],  # record_id
                            log[2],  # employee_id
                            emp_name,
                            log[3] if log[3] else "N/A",  # terminal_id
                            log[4],  # timestamp
                            log[5],  # type (IN/OUT)
                            log[6],  # method
                            log[7] if log[7] else "N/A",  # confidence
                            log[8] if log[8] else "ACCEPTED",  # status
                            log[12] if log[12] else "N/A",  # photo_path
                            log[13]  # integrity_hash
                        ])
                
                QMessageBox.information(self, "Succès", f"Journaux exportés vers {filename}")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Échec d'exportation des journaux : {str(e)}")

    def logout(self):
        self.close()
        self.public.show()