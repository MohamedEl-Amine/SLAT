"""
Database module for SLAT using SQLite.
Handles all database operations.
"""

import sqlite3
import os
from datetime import datetime
from cryptography.fernet import Fernet
import hashlib
from models import Employee, AttendanceRecord

class Employee:
    def __init__(self, id, employee_id, name, enabled, qr_code, face_embedding, created_at):
        self.id = id
        self.employee_id = employee_id
        self.name = name
        self.enabled = enabled
        self.qr_code = qr_code
        self.face_embedding = face_embedding
        self.created_at = created_at

class Database:
    def __init__(self, db_path="data/slat.db"):
        self.db_path = db_path
        self.key_path = "data/key.key"
        self._ensure_data_dir()
        self.cipher = self._load_or_create_key()
        self._create_tables()
        self._migrate_database()

    def _ensure_data_dir(self):
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)

    def _load_or_create_key(self):
        if os.path.exists(self.key_path):
            with open(self.key_path, 'rb') as f:
                key = f.read()
        else:
            key = Fernet.generate_key()
            with open(self.key_path, 'wb') as f:
                f.write(key)
        return Fernet(key)

    def _create_tables(self):
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()

            # Employees table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS employees (
                    id INTEGER PRIMARY KEY,
                    employee_id TEXT UNIQUE NOT NULL,
                    name TEXT NOT NULL,
                    enabled BOOLEAN DEFAULT 1,
                    qr_code TEXT,
                    face_embedding BLOB,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            # Attendance logs table (append-only audit trail)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS attendance_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    record_id TEXT UNIQUE NOT NULL,
                    employee_id TEXT NOT NULL,
                    terminal_id TEXT NOT NULL,
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    type TEXT NOT NULL,  -- 'IN' or 'OUT'
                    method TEXT NOT NULL,  -- 'FACE', 'CARD', 'QR', 'MANUAL'
                    confidence REAL,  -- Recognition confidence (if face)
                    status TEXT DEFAULT 'ACCEPTED',  -- 'ACCEPTED' or 'REJECTED'
                    operator_id TEXT,  -- Admin who corrected it (if any)
                    correction_reason TEXT,  -- Why corrected
                    replaces_record_id TEXT,  -- If this corrects another record
                    photo_path TEXT,  -- Path to checkpoint photo
                    integrity_hash TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    modified_at TIMESTAMP,
                    FOREIGN KEY (employee_id) REFERENCES employees(employee_id)
                )
            ''')

            # Payroll summaries table (aggregated data)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS payroll_summaries (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id TEXT NOT NULL,
                    date DATE NOT NULL,
                    first_in TIME,
                    last_out TIME,
                    total_hours REAL,
                    overtime REAL DEFAULT 0,
                    late_minutes INTEGER DEFAULT 0,
                    early_leave_minutes INTEGER DEFAULT 0,
                    status TEXT DEFAULT 'NORMAL',  -- 'NORMAL', 'ABSENT', 'EXCEPTION'
                    generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(employee_id, date),
                    FOREIGN KEY (employee_id) REFERENCES employees(employee_id)
                )
            ''')

            # Settings table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS settings (
                    key TEXT PRIMARY KEY,
                    value TEXT
                )
            ''')

            # Insert default settings
            default_settings = [
                ('morning_start', '08:30'),
                ('morning_end', '09:00'),
                ('afternoon_start', '16:00'),
                ('afternoon_end', '16:30'),
                ('admin_password', hashlib.sha256('admin'.encode()).hexdigest()),
                ('card_enabled', '1'),
                ('qr_enabled', '0'),
                ('face_enabled', '0'),
                ('attendance_mode', 'qr')  # qr, face, or card
            ]

            for key, value in default_settings:
                cursor.execute('INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)', (key, value))

            conn.commit()

    def _migrate_database(self):
        """Migrate database schema for face recognition compliance"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()

            # Check if face_image column exists and face_embedding doesn't
            cursor.execute("PRAGMA table_info(employees)")
            columns = cursor.fetchall()
            column_names = [col[1] for col in columns]

            if 'face_image' in column_names and 'face_embedding' not in column_names:
                print("Migrating database: face_image -> face_embedding")

                # Rename face_image to face_embedding
                cursor.execute("ALTER TABLE employees RENAME COLUMN face_image TO face_embedding")

                # Note: In a real migration, you would need to convert existing face images to embeddings
                # For now, we'll clear the face data since raw images are not allowed per specs
                cursor.execute("UPDATE employees SET face_embedding = NULL WHERE face_embedding IS NOT NULL")

                conn.commit()
                print("Database migration completed")

    def get_setting(self, key):
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT value FROM settings WHERE key = ?', (key,))
            result = cursor.fetchone()
            return result[0] if result else None

    def update_setting(self, key, value):
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', (key, value))
            conn.commit()

    def get_employee(self, employee_id):
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM employees WHERE employee_id = ?', (employee_id,))
            row = cursor.fetchone()
            if row:
                created_at = datetime.fromisoformat(row[6]) if row[6] else None
                return Employee(id=row[0], employee_id=row[1], name=row[2], enabled=bool(row[3]), qr_code=row[4], face_embedding=row[5], created_at=created_at)
            return None
    
    def get_employee_by_qr(self, qr_code):
        """Get employee by QR code"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM employees WHERE qr_code = ?', (qr_code,))
            row = cursor.fetchone()
            if row:
                created_at = datetime.fromisoformat(row[6]) if row[6] else None
                return Employee(id=row[0], employee_id=row[1], name=row[2], enabled=bool(row[3]), qr_code=row[4], face_embedding=row[5], created_at=created_at)
            return None

    def record_attendance(self, employee_id, action, method_used, device_id, photo_path=None, confidence=None, operator_id=None):
        """Record attendance with full audit trail"""
        import uuid
        timestamp = datetime.now()
        record_id = str(uuid.uuid4())
        
        # Create integrity hash
        hash_input = f"{record_id}{employee_id}{action}{timestamp}{method_used}".encode()
        integrity_hash = hashlib.sha256(hash_input).hexdigest()

        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO attendance_logs (
                    record_id, employee_id, terminal_id, timestamp, type, method, 
                    confidence, status, operator_id, photo_path, integrity_hash, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (record_id, employee_id, device_id, timestamp, action, method_used, 
                  confidence, 'ACCEPTED', operator_id, photo_path, integrity_hash, timestamp))
            conn.commit()
            return record_id

    def correct_attendance(self, original_record_id, operator_id, correction_reason, new_type=None, new_timestamp=None):
        """Correct an attendance record by creating a new entry (immutable audit trail)"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Get original record
            cursor.execute('SELECT * FROM attendance_logs WHERE record_id = ?', (original_record_id,))
            original = cursor.fetchone()
            
            if not original:
                return None
            
            # Mark original as corrected
            cursor.execute('''
                UPDATE attendance_logs 
                SET status = 'CORRECTED', modified_at = ? 
                WHERE record_id = ?
            ''', (datetime.now(), original_record_id))
            
            # Create new corrected record
            import uuid
            new_record_id = str(uuid.uuid4())
            timestamp = new_timestamp if new_timestamp else original[4]
            action = new_type if new_type else original[5]
            
            hash_input = f"{new_record_id}{original[2]}{action}{timestamp}{original[6]}".encode()
            integrity_hash = hashlib.sha256(hash_input).hexdigest()
            
            cursor.execute('''
                INSERT INTO attendance_logs (
                    record_id, employee_id, terminal_id, timestamp, type, method,
                    confidence, status, operator_id, correction_reason, replaces_record_id,
                    photo_path, integrity_hash, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (new_record_id, original[2], original[3], timestamp, action, original[6],
                  original[7], 'ACCEPTED', operator_id, correction_reason, original_record_id,
                  original[12], integrity_hash, datetime.now()))
            
            conn.commit()
            return new_record_id

    def add_employee(self, employee_id, name, qr_code=None, face_embedding=None):
        """Add a new employee to the database"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            try:
                cursor.execute('''
                    INSERT INTO employees (employee_id, name, enabled, qr_code, face_embedding)
                    VALUES (?, ?, 1, ?, ?)
                ''', (employee_id, name, qr_code, face_embedding))
                conn.commit()
                return True
            except sqlite3.IntegrityError:
                return False
    
    def update_employee_qr(self, employee_id, qr_code):
        """Update employee QR code"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE employees SET qr_code = ? WHERE employee_id = ?', (qr_code, employee_id))
            conn.commit()
    
    def update_employee_face(self, employee_id, face_embedding):
        """Update employee face embedding"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE employees SET face_embedding = ? WHERE employee_id = ?', (face_embedding, employee_id))
            conn.commit()
    
    def generate_qr_code(self, employee_id):
        """Generate QR code for employee"""
        import qrcode
        from io import BytesIO
        
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(employee_id)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Convert to bytes
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        qr_bytes = buffer.getvalue()
        
        # Update employee with QR code data
        self.update_employee_qr(employee_id, employee_id)  # Store employee_id as qr_code
        
        return qr_bytes

    def get_all_employees(self):
        """Get all employees"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM employees ORDER BY name')
            rows = cursor.fetchall()
            employees = []
            for row in rows:
                created_at = datetime.fromisoformat(row[6]) if row[6] else None
                employees.append(Employee(
                    id=row[0], 
                    employee_id=row[1], 
                    name=row[2], 
                    enabled=bool(row[3]), 
                    qr_code=row[4], 
                    face_embedding=row[5], 
                    created_at=created_at
                ))
            return employees

    def update_employee_status(self, employee_id, enabled):
        """Enable or disable an employee"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE employees SET enabled = ? WHERE employee_id = ?', (int(enabled), employee_id))
            conn.commit()

    def update_employee_name(self, employee_id, new_name):
        """Update employee name"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE employees SET name = ? WHERE employee_id = ?', (new_name, employee_id))
            conn.commit()
            return cursor.rowcount > 0

    def get_all_logs(self, limit=None):
        """Get all attendance logs"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            query = 'SELECT * FROM attendance_logs ORDER BY timestamp DESC'
            if limit:
                query += f' LIMIT {limit}'
            cursor.execute(query)
            return cursor.fetchall()

    def get_employee_logs(self, employee_id, limit=None):
        """Get attendance logs for specific employee"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            query = 'SELECT * FROM attendance_logs WHERE employee_id = ? ORDER BY timestamp DESC'
            if limit:
                query += f' LIMIT {limit}'
            cursor.execute(query, (employee_id,))
            rows = cursor.fetchall()
            
            # Convert to AttendanceRecord objects
            records = []
            for row in rows:
                records.append(AttendanceRecord(
                    id=row[0],
                    employee_id=row[2],
                    action=row[5],
                    timestamp=datetime.fromisoformat(row[4]) if isinstance(row[4], str) else row[4],
                    method_used=row[6],
                    device_id=row[3],
                    photo=None,
                    integrity_hash=row[13]
                ))
            return records

    def generate_payroll_summary(self, start_date, end_date):
        """Generate payroll summary for date range"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Get all employees
            employees = self.get_all_employees()
            
            summaries = []
            for emp in employees:
                # Get attendance logs for employee in date range
                cursor.execute('''
                    SELECT DATE(timestamp), type, TIME(timestamp)
                    FROM attendance_logs 
                    WHERE employee_id = ? 
                    AND DATE(timestamp) BETWEEN ? AND ?
                    AND status = 'ACCEPTED'
                    ORDER BY timestamp
                ''', (emp.employee_id, start_date, end_date))
                
                logs = cursor.fetchall()
                
                # Group by date
                daily_logs = {}
                for log in logs:
                    date = log[0]
                    if date not in daily_logs:
                        daily_logs[date] = {'in': [], 'out': []}
                    
                    if log[1] == 'IN':
                        daily_logs[date]['in'].append(log[2])
                    else:
                        daily_logs[date]['out'].append(log[2])
                
                # Calculate daily summaries
                morning_start = datetime.strptime(self.get_setting('morning_start'), '%H:%M').time()
                
                for date, logs in daily_logs.items():
                    first_in = min(logs['in']) if logs['in'] else None
                    last_out = max(logs['out']) if logs['out'] else None
                    
                    # Calculate total hours
                    total_hours = 0
                    if first_in and last_out:
                        in_dt = datetime.combine(datetime.strptime(date, '%Y-%m-%d').date(), 
                                                datetime.strptime(first_in, '%H:%M:%S').time())
                        out_dt = datetime.combine(datetime.strptime(date, '%Y-%m-%d').date(),
                                                 datetime.strptime(last_out, '%H:%M:%S').time())
                        total_hours = (out_dt - in_dt).total_seconds() / 3600
                    
                    # Calculate lateness
                    late_minutes = 0
                    if first_in:
                        in_time = datetime.strptime(first_in, '%H:%M:%S').time()
                        if in_time > morning_start:
                            late_minutes = int((datetime.combine(datetime.min, in_time) - 
                                              datetime.combine(datetime.min, morning_start)).total_seconds() / 60)
                    
                    # Calculate overtime (hours > 8)
                    overtime = max(0, total_hours - 8)
                    
                    status = 'NORMAL' if first_in and last_out else 'EXCEPTION'
                    
                    summaries.append({
                        'employee_id': emp.employee_id,
                        'employee_name': emp.name,
                        'date': date,
                        'first_in': first_in,
                        'last_out': last_out,
                        'total_hours': round(total_hours, 2),
                        'overtime': round(overtime, 2),
                        'late_minutes': late_minutes,
                        'status': status
                    })
            
            return summaries
    
    def export_payroll_csv(self, start_date, end_date, filepath):
        """Export payroll summary to CSV"""
        import csv
        
        summaries = self.generate_payroll_summary(start_date, end_date)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=[
                'employee_id', 'employee_name', 'date', 'first_in', 'last_out', 
                'total_hours', 'overtime', 'late_minutes', 'status'
            ])
            writer.writeheader()
            writer.writerows(summaries)
        
        return filepath
    
    def export_payroll_excel(self, start_date, end_date, filepath):
        """Export payroll summary to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            # Fallback to CSV if openpyxl not available
            return self.export_payroll_csv(start_date, end_date, filepath.replace('.xlsx', '.csv'))
        
        summaries = self.generate_payroll_summary(start_date, end_date)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payroll Summary"
        
        # Headers
        headers = ['Employee ID', 'Name', 'Date', 'First In', 'Last Out', 
                   'Total Hours', 'Overtime', 'Late (min)', 'Status']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for summary in summaries:
            ws.append([
                summary['employee_id'],
                summary['employee_name'],
                summary['date'],
                summary['first_in'],
                summary['last_out'],
                summary['total_hours'],
                summary['overtime'],
                summary['late_minutes'],
                summary['status']
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        return filepath
    
    def export_audit_trail_csv(self, start_date, end_date, filepath):
        """Export full audit trail to CSV"""
        import csv
        
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT record_id, employee_id, terminal_id, timestamp, type, method,
                       confidence, status, operator_id, correction_reason, replaces_record_id,
                       created_at, modified_at
                FROM attendance_logs
                WHERE DATE(timestamp) BETWEEN ? AND ?
                ORDER BY timestamp DESC
            ''', (start_date, end_date))
            
            rows = cursor.fetchall()
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Record ID', 'Employee ID', 'Terminal ID', 'Timestamp', 'Type', 
                           'Method', 'Confidence', 'Status', 'Operator ID', 'Correction Reason',
                           'Replaces Record ID', 'Created At', 'Modified At'])
            writer.writerows(rows)
        
        return filepath

    def hash_password(self, password):
        """Hash a password"""
        return hashlib.sha256(password.encode()).hexdigest()

    def verify_password(self, password, hashed):
        """Verify a password against a hash"""
        return hashlib.sha256(password.encode()).hexdigest() == hashed